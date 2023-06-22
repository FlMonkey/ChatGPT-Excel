from openpyxl import load_workbook
import os
import openai

openai.organization = 'org...' #org id
openai.api_key = 'sk-...' #secret key
wb = load_workbook(filename='filehere')
ws = wb.active

nr = input("collum letter: ")
nmin = int(input("first row number: "))
nmax = int(input("last row number: "))

for i in range(nmin, nmax):
    current = str(ws[nr + str(i)].value)
    
    messages= [ {"role": "system", "content" : "prompt for gpt here " + current}]
    
    completion = openai.ChatCompletion.create(
        model="gpt-4",
        messages=messages
    )
    chat_response = completion.choices[0].message.content
    pr = chr(int(ord(nr))-1)
    print(pr)
    ws[pr + str(i)] = chat_response
    print(pr + str(i))
    print(f'progress: {(i-nmin)+1}/{nmax-nmin}')
    
wb.save(filename='file name here')
print("Saved to file")
    #print(print(f'\nChatGPT: {chat_response}\n'))

    #print(current)
